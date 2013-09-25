# -*- coding: latin-1 -*-

# Python imports
import warnings, codecs, textwrap,os, itertools, difflib, zipfile, types
from multiprocessing import Process

# wxPython imports
import wx, wx.grid, wx.stc
from wx.lib.mixins.listctrl import CheckListCtrlMixin,TextEditMixin,ListCtrlAutoWidthMixin

import CoolProp
from CoolProp.State import State
from CoolProp import CoolProp as CP

import numpy as np

import matplotlib as mpl
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as WXCanvas
from matplotlib.backends.backend_wxagg import NavigationToolbar2Wx as WXToolbar

from PDSim.scroll import scroll_geo
from PDSim.scroll.plots import plotScrollSet, ScrollAnimForm
from PDSim.misc.datatypes import AnnotatedValue
from datatypes import AnnotatedGUIObject, HeaderStaticText, CoupledAnnotatedGUIObject

import PDSimGUI

import h5py
import quantities as pq
from time import clock

length_units = {
                'Meter': pq.length.m,
                'Millimeter': pq.length.mm,
                'Micrometer' : pq.length.um,
                'Centimeter' : pq.length.cm,
                'Inch' : pq.length.inch,
                }

area_units = {
                'Square Meter': pq.length.m**2,
                'Square Micrometer' : pq.length.um**2,
                'Square Centimeter' : pq.length.cm**2,
                'Square Inch' : pq.length.inch**2,
                }

volume_units = {
                'Cubic Meter': pq.length.m**3,
                'Cubic Micrometer' : pq.length.um**3,
                'Cubic Centimeter' : pq.length.cm**3,
                'Cubic Inch' : pq.length.inch**3,
                }

pressure_units = {
                  'kPa' : pq.kPa,
                  'psia' : pq.psi
                  }
rev = pq.UnitQuantity('revolution', 2*np.pi*pq.radians, symbol='rev')

rotational_speed_units ={
                         'Radians per second': pq.radians/pq.sec,
                         'Radians per minute': pq.radians/pq.min,
                         'Revolutions per second': rev/pq.sec,
                         'Revolutions per minute': rev/pq.min,
                         }

temperature_units = {
                     'Kelvin' : np.nan,
                     'Celsius' : np.nan,
                     'Fahrenheit' : np.nan,
                     'Rankine': np.nan
                     }


class InputsToolBook(wx.Toolbook):
    
    def get_script_chunks(self):
        """
        Pull all the values out of the child panels, using the values in 
        self.items and the function get_script_chunks if the panel implements
        it
        
        The values are written into the script file that will be execfile-d
        """
        chunks = []
        for panel in self.panels:
            chunks.append('#############\n# From '+panel.Name+'\n############\n')
            if hasattr(panel,'get_script_params'):
                chunks.append(panel.get_script_params())
            if hasattr(panel,'get_script_chunks'):
                chunks.append(panel.get_script_chunks())
        return chunks

class UnitConvertor(wx.Dialog):
    def __init__(self, value, default_units, type = None, TextCtrl = None):
        wx.Dialog.__init__(self, None, title='Convert units')
        
        self.default_units = default_units
        self.__is_temperature__ = False
        if default_units in length_units or type == 'length':
            self.unit_dict = length_units
        elif default_units in area_units or type == 'area':
            self.unit_dict = area_units
        elif default_units in volume_units or type == 'volume':
            self.unit_dict = volume_units
        elif default_units in rotational_speed_units or type == 'rotational_speed':
            self.unit_dict = rotational_speed_units
        elif default_units in pressure_units or type == 'pressure':
            self.unit_dict = pressure_units
        elif default_units in temperature_units or type == 'temperature':
            self.unit_dict = temperature_units
            self.__is_temperature__ = True
        else:
            raise KeyError('Sorry your units '+default_units+' did not match any of the unit terms')
            
        self.txt = wx.TextCtrl(self, value=str(value))
        self.units = wx.Choice(self)
        self.units.AppendItems(sorted(self.unit_dict.keys()))
        if default_units in self.units.GetStrings():
            self.units.SetStringSelection(default_units)
        else:
            raise KeyError
        sizer = wx.BoxSizer(wx.HORIZONTAL)
        sizer.Add(self.txt)
        sizer.Add(self.units)
        self.SetSizer(sizer)
        self.Fit()
        self._old_units = default_units
        
        self.Bind(wx.EVT_CHOICE, self.OnSwitchUnits, self.units)

    #Bind a key-press event to all objects to get Esc 
        children = self.GetChildren()
        for child in children:
            child.Bind(wx.EVT_KEY_UP,  self.OnKeyPress)
        
    def OnKeyPress(self,event=None):
        """ cancel if Escape key is pressed """
        event.Skip()
        if event.GetKeyCode() == wx.WXK_ESCAPE:
            self.EndModal(wx.ID_CANCEL)
        elif event.GetKeyCode() == wx.WXK_RETURN:
            self.EndModal(wx.ID_OK)
                    
    def OnSwitchUnits(self, event):
        if not self.__is_temperature__:
            old = float(self.txt.GetValue()) * self.unit_dict[self._old_units]
            new_units_str = self.units.GetStringSelection()
            new_units = self.unit_dict[new_units_str]
            old.units = new_units
            self._old_units = new_units_str
            self.txt.SetValue(str(old.magnitude))
        else:
            old_val = float(self.txt.GetValue())
            new_units_str = self.units.GetStringSelection()
            new_val = self._temperature_convert(self._old_units, old_val, new_units_str)
            self._old_units = new_units_str
            self.txt.SetValue(str(new_val))
            
    def get_value(self):
        """
        Return a string with value in the original units for use in calling function to dialog
        """
        if not self.__is_temperature__:
            old = float(self.txt.GetValue()) * self.unit_dict[self._old_units]
            old.units = self.unit_dict[self.default_units]
            return str(old.magnitude)
        else:
            old_val = float(self.txt.GetValue())
            return str(self._temperature_convert(self._old_units, old_val, self.default_units))
    
    def _temperature_convert(self, old, old_val, new):
        """
        Internal method to convert temperature
        
        Parameters
        ----------
        old : string
        old_val : float
        new : string
        """
        #convert old value to Celsius
        # also see: http://en.wikipedia.org/wiki/Temperature
        if old == 'Fahrenheit':
            celsius_val = (old_val-32)*5.0/9.0
        elif old == 'Kelvin':
            celsius_val = old_val-273.15
        elif old == 'Rankine':
            celsius_val = (old_val-491.67)*5.0/9.0
        elif old == 'Celsius':
            celsius_val = old_val
            
        #convert celsius to new value
        if new == 'Celsius':
            return celsius_val
        elif new == 'Fahrenheit':
            return celsius_val*9.0/5.0+32.0
        elif new == 'Kelvin':
            return celsius_val+273.15
        elif new == 'Rankine':
            return (celsius_val+273.15)*9.0/5.0
        
def mathtext_to_wxbitmap(s):
    #The magic from http://matplotlib.org/examples/user_interfaces/mathtext_wx.html?highlight=button
    from matplotlib.mathtext import MathTextParser
    
    mathtext_parser = MathTextParser("Bitmap")
    ftimage, depth = mathtext_parser.parse(s, 100)
    return wx.BitmapFromBufferRGBA(ftimage.get_width(), 
                                   ftimage.get_height(),
                                   ftimage.as_rgba_str()
                                   )
        
def EquationButtonMaker(LaTeX, parent, **kwargs):
    """
    A Convenience function to generate a button with LaTeX as its image
    
    LaTeX : string
    parent : wx.Window
    
    kwargs passed to BitmapButton constructor 
    """
    return wx.BitmapButton(parent, bitmap = mathtext_to_wxbitmap(LaTeX), **kwargs)
    
def LaTeXImageMaker(LaTeX,parent,**kwargs):
    return wx.StaticBitmap(parent, bitmap = mathtext_to_wxbitmap(LaTeX), **kwargs)        
        
class PlotPanel(wx.Panel):
    def __init__(self, parent, toolbar = False, **kwargs):
        wx.Panel.__init__(self, parent, **kwargs)
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.figure = mpl.figure.Figure(dpi=100, figsize=(2, 2))
        self.canvas = WXCanvas(self, -1, self.figure)
        
        sizer.Add(self.canvas)
        
        if toolbar:
            self.toolbar = WXToolbar(self.canvas)
            self.toolbar.Realize()
            sizer.Add(self.toolbar)
            
        self.SetSizer(sizer)
        sizer.Layout()

class PDPanel(wx.Panel):
    """
    A base class for panel with some goodies thrown in 
    
    Not intended for direct instantiation, rather it should be 
    subclassed, like in :class:`recip_panels.GeometryPanel`
    
    Loading from configuration file
    -------------------------------
    Method (A)
    
    Any subclassed PDPanel must either store a list of items as 
    ``self.items`` where the entries are dictionaries with at least the entries:
    
    - ``text``: The text description of the term 
    - ``attr``: The attribute in the simulation class that is linked with this term
    
    Other terms that can be included in the item are:
    - ``tooltip``: The tooltip to be attached to the textbox
    
    Method (B)
    
    The subclassed panel can provide the functions post_prep_for_configfile
    and post_get_from_configfile, each of which take no inputs.
    
    In post_prep_for_configfile, the subclassed panel can package up its elements
    into a form that can then be re-created when post_get_from_configfile is
    called
    
    
    Saving to configuration file
    ----------------------------
    
    Adding terms to parametric table
    --------------------------------
    
    """
    def __init__(self,*args,**kwargs):
        wx.Panel.__init__(self,*args,**kwargs)
        self.name=kwargs.get('name','')
        
        # Get the main frame
        self.main = self.GetTopLevelParent()
        
    def _get_item_by_attr(self, attr):
        if hasattr(self,'items'):
            for item in self.items:
                if item['attr'] == attr:
                    return item
        raise ValueError('_get_item_by_attr failed')
        
        
    def get_annotated_values(self, keys, config = None):
        
        if config is None:
            if hasattr(self,'config'):
                config = self.config
            else:
                config = dict()
        
        annotated_values = []
        for key in keys:
            mapped_val = self.desc_map[key]
            if len(mapped_val) == 2:
                # Get the annotation and the units for the term (no default provided)
                annotation, units = mapped_val 
                # Add the annotated object to the list of objects
                annotated_values.append(AnnotatedValue(key, config[key], annotation, units))
            elif len(mapped_val) == 3:
                # Get the annotation and the units for the term 
                annotation, units, default = mapped_val 
                if key in config:
                    # Add the annotated object to the list of objects
                    annotated_values.append(AnnotatedValue(key, config[key], annotation, units))
                else:
                    # Add the annotated object to the list of objects
                    annotated_values.append(AnnotatedValue(key, default, annotation, units))
                    
            self.keys_for_config.append(key)
        return annotated_values
    
    def _get_value(self,thing):
        #This first should work for wx.TextCtrl
        if hasattr(thing,'GetValue'):
            value=str(thing.GetValue()).strip()
            try:
                return float(value)
            except ValueError:
                return value
        elif hasattr(thing,'GetSelectionString'):
            value=thing.GetSelection()
            try:
                return float(value)
            except ValueError:
                return value
             
    def get_script_params(self):
        if not hasattr(self,'items'):
            return ''
        else:
            items = self.items
        
        if hasattr(self,'skip_list'):
            # Don't actually set these attributes (they might over-write 
            # methods or attributes in the simulation)
            items = [item for item in items if item['attr'] not in self.skip_list()]
        
        values = ''
        for item in items:
            values += 'sim.{attr:s} = {val:s}\n'.format(attr = item['attr'],
                                                        val = str(self._get_value(item['textbox'])))
        
        return values
    
    def ConstructItems(self, items, sizer, configdict = None, descdict=None, parent = None):
        
        """
        Parameters
        ----------
        items : a list of dictionaries
            Each item is a dictionary of values with the keys:
                attr : PDSim attribute
        sizer : wx.Sizer
            The sizer to add the items to
        configdict : dictionary
            The configuration value dictionary that was pulled from the file
        descdict : dictionary
            The configuration description dictionary that was pulled from the file
        parent : wx.Window
            The parent of the items, default is this panel
        """
        for item in items:
            
            if parent is None:
                parent = self
            
            if 'val' not in item and configdict is not None:
                k = item['attr']
                if k not in configdict:
                    #Returns a dictionary of values and a dictionary of descriptions
                    d,desc = self.get_from_configfile(self.name, k, default = True)
                    #Get the entries for the given key
                    val,item['text'] = d[k],desc[k]
                else:
                    val = configdict[k]
                    item['text'] = descdict[k]
            else:
                d,desc = self.get_from_configfile(self.name, item['attr'])
                #Get the entries for the given key
                val,item['text'] = d[k],desc[k]
                
            label=wx.StaticText(parent, -1, item['text'])
            sizer.Add(label, 1, wx.EXPAND)
            textbox=wx.TextCtrl(parent, -1, str(val))
            sizer.Add(textbox, 1, wx.EXPAND)
            item.update(dict(textbox=textbox, label=label))
            
            caption = item['text']
            if caption.find(']')>=0 and caption.find(']')>=0: 
                units = caption.split('[',1)[1].split(']',1)[0]
                unicode_units = unicode(units)
                if unicode_units == u'm':
                    textbox.default_units = 'Meter'
                elif unicode_units == u'm\xb2':
                    textbox.default_units = 'Square Meter'
                elif unicode_units == u'm\xb3':
                    textbox.default_units = 'Cubic Meter'
                elif units == 'rad/s':
                    textbox.default_units = 'Radians per second'
                self.Bind(wx.EVT_CONTEXT_MENU,self.OnChangeUnits,textbox)     
                
    def construct_items(self, annotated_objects, sizer = None, parent = None):
        
        """
        Parameters
        ----------
        annotated_objects : a list of `AnnotatedValue <PDSim.misc.datatypes.AnnotatedValue>`
        sizer : wx.Sizer
            The sizer to add the wx elements to
        parent : wx.Window
            The parent of the items, default is this panel
            
        Returns
        -------
        annotated_GUI_objects : a list of `GUIAnnotatedObject` derived from the input ``annotated_objects`` 
        """
        
        # Default to parent it to this panel
        if parent is None:
            parent = self
            
        # Output list of annotated GUI objects
        annotated_GUI_objects = []
        
        # Loop over the objects
        for o in annotated_objects:
            
            # Type-check
            if not isinstance(o, AnnotatedValue):
                raise TypeError('object of type [{t:s}] is not an AnnotatedValue'.format(t = type(o)))
                
            # Build the GUI objects
            label=wx.StaticText(parent, -1, o.annotation)
            
            if sizer is not None:
                # Add the label to the sizer
                sizer.Add(label, 1, wx.EXPAND)

            # If the input is a boolean value, make a check box
            if isinstance(o.value,bool):
                # Build the checkbox
                checkbox = wx.CheckBox(parent)
                # Set its value
                checkbox.SetValue(o.value)
                
                if sizer is not None:
                    # Add to the sizer
                    sizer.Add(checkbox, 1, wx.EXPAND)
                
                # Add to the annotated objects
                annotated_GUI_objects.append(AnnotatedGUIObject(o, checkbox))
            
            # Otherwise make a TextCtrl 
            else:
                # Create the textbox
                textbox=wx.TextCtrl(parent, value = str(o.value))
                
                if sizer is not None:
                    # Add it to the sizer
                    sizer.Add(textbox, 1, wx.EXPAND)
                
                # Units are defined for the item
                if o.units:
                    unicode_units = unicode(o.units)
                    textbox.default_units = ''
                    if unicode_units == u'm':
                        textbox.default_units = 'Meter'
                    elif unicode_units == u'm^2':
                        textbox.default_units = 'Square Meter'
                    elif unicode_units == u'm^3':
                        textbox.default_units = 'Cubic Meter'
                    elif unicode_units == u'rad/s':
                        textbox.default_units = 'Radians per second'
                    elif unicode_units == u'kPa':
                        textbox.default_units = 'kPa'
                    elif unicode_units == u'K':
                        textbox.default_units = 'Kelvin'
                    
                    #If it has units bind the unit changing callback on right-click
                    if textbox.default_units:
                        self.Bind(wx.EVT_CONTEXT_MENU,self.OnChangeUnits,textbox)  
                
                annotated_GUI_objects.append(AnnotatedGUIObject(o,textbox))
            
        if len(annotated_GUI_objects) == 1:
            return annotated_GUI_objects[0]
        else:
            return annotated_GUI_objects
    
    def BindChangeUnits(self, TextCtrl):
        self.Bind(wx.EVT_KEY_DOWN, self.OnChangeUnits, TextCtrl)
        
    def OnChangeUnits(self, event):
        TextCtrl = event.GetEventObject()
        dlg = UnitConvertor(value = float(TextCtrl.GetValue()),
                            default_units = TextCtrl.default_units
                            )
        
        dlg.ShowModal()
        TextCtrl.SetValue(dlg.get_value())
        dlg.Destroy()
        
class OutputTreePanel(wx.Panel):
    
    def __init__(self, parent, runs):
        
        import wx, textwrap
        import wx.gizmos
        from operator import mul
        
        wx.Panel.__init__(self, parent, -1)
        self.Bind(wx.EVT_SIZE, self.OnSize)
        
        self.tree = wx.gizmos.TreeListCtrl(self, -1, style =
                                           wx.TR_DEFAULT_STYLE
                                           #| wx.TR_HAS_BUTTONS
                                           #| wx.TR_TWIST_BUTTONS
                                           #| wx.TR_ROW_LINES
                                           #| wx.TR_COLUMN_LINES
                                           #| wx.TR_NO_LINES 
                                           | wx.TR_FULL_ROW_HIGHLIGHT
                                           )
        
#        # Make a list of list of keys for each HDF5 file
#        lists_of_keys = []
#        for run in runs:
#            keys = []
#            run.visit(keys.append) #For each thing in HDF5 file, append its name to the keys list
#            lists_of_keys.append(keys)
#        
#        from time import clock
#        t1 = clock()
#        
#        mylist = []
#        def func(name, obj):
#            if isinstance(obj, h5py.Dataset):
#                try:
#                    mylist.append((name, obj.value))
#                except ValueError:
#                    pass
#            elif isinstance(obj, h5py.Group):
#                mylist.append((name, None))
#            else:
#                return
#                
#        run.visititems(func)
#        
#        for el in mylist:
#            r,v = el
#            
#            # Always make this level
#            child = self.tree.AppendItem(self.root, r)
#            self.tree.SetItemImage(child, fldridx, which = wx.TreeItemIcon_Normal)
#            self.tree.SetItemImage(child, fldropenidx, which = wx.TreeItemIcon_Expanded)
#            self.tree.SetItemText(child, str(v), 1)
#            
#        t2 = clock()
#        print t2 - t1, 'to get all the items'
            
        self.runs = runs
        self.Disable()
        self.rebuild()
        
    def rebuild(self):
        
        self.Enable()
        
        # Remove all the columns that are in the tree
        ncols = self.tree.GetColumnCount()
        if ncols > 1: #There's already something there
            for col in range(0, ncols):
                self.tree.RemoveColumn(0) # Keep removing the first column
             
        self.tree.DeleteAllItems()
        self.tree.DeleteRoot()

        # If all the runs that are being run have the run_index key, sort based on it
        if self.runs and all(['run_index' in run.keys() for run in self.runs]):
            s = sorted([(run['run_index'].value,run) for run in self.runs])
            indices, runs = zip(*s)
            self.runs = list(runs)
        
        if not self.runs:
            return
        
        isz = (16, 16)
        il = wx.ImageList(*isz)
        fldridx     = il.Add(wx.ArtProvider_GetBitmap(wx.ART_FOLDER,      wx.ART_OTHER, isz))
        fldropenidx = il.Add(wx.ArtProvider_GetBitmap(wx.ART_FILE_OPEN,   wx.ART_OTHER, isz))
        fileidx     = il.Add(wx.ArtProvider_GetBitmap(wx.ART_NORMAL_FILE, wx.ART_OTHER, isz))
        
        self.tree.SetImageList(il)
        self.il = il
        self.tree.AddColumn("Main column")
        self.tree.SetMainColumn(0) # the one with the tree in it...
        self.tree.SetColumnWidth(0, 175)
        
        # Build the columns
        for i, run in enumerate(self.runs):
            self.tree.AddColumn("Run {i:d}".format(i = i))  

        # Create the root column
        self.root = self.tree.AddRoot("The Root Item")
        
        self.tree.SetItemImage(self.root, fldridx, which = wx.TreeItemIcon_Normal)
        self.tree.SetItemImage(self.root, fldropenidx, which = wx.TreeItemIcon_Expanded)
        
        def _recursive_hdf5_add(root, objects):
            for thing in objects[0]:
                # Always make this level
                child = self.tree.AppendItem(root, str(thing))
                
                # If it is a dataset, write the dataset contents to the tree
                if isinstance(objects[0][thing], h5py.Dataset):
                    for i, o in enumerate(objects):
                        if not o[thing].shape: # It's a single element, perhaps a number or a string.  
                                               # shape will be an empty tuple, hence not () is True
                            self.tree.SetItemText(child, str(o[thing].value), i+1)
                        else:
                            # A place holder for now - will develop a frame to display the matrix
                            self.tree.SetItemText(child, str(o[thing]), i+1)
                        
                        if o[thing].attrs and 'note' in o[thing].attrs: #If it has any annotations
                            self.tree.SetItemPyData(child, dict(annotation = o[thing].attrs['note']))
                            self.tree.SetItemBackgroundColour(child,(255,0,0)) #Color the cell
                
                # Otherwise if it is a group, change the icon and recurse into the group
                elif isinstance(objects[0][thing], h5py.Group):
                    self.tree.SetItemImage(child, fldridx, which = wx.TreeItemIcon_Normal)
                    self.tree.SetItemImage(child, fldropenidx, which = wx.TreeItemIcon_Expanded)
                    try:
                        _recursive_hdf5_add(child, [o[thing] for o in objects])
                    except KeyError as KE:
                        print KE
        
        t1 = clock()
        _recursive_hdf5_add(self.root, self.runs)
        t2 = clock()
        
        print t2-t1,'secs elapsed to load output tree'
        
        self.tree.Expand(self.root)

        self.tree.GetMainWindow().Bind(wx.EVT_RIGHT_UP, self.OnRightUp)
        self.tree.Bind(wx.EVT_TREE_SEL_CHANGED, self.OnActivate)
        
    def OnSaveXLSX(self):
        
        def get_row(item):
            return [self.tree.GetItemText(item,col) for col in range(self.tree.GetColumnCount())]
        
        def get_path(item):
            parents = []
            counter = 0
            while True:
                item = self.tree.GetItemParent(item)
                if not item or item == self.tree.GetRootItem():
                    parents.reverse()
                    return parents
                name = self.tree.GetItemText(item,0)
                parents.append(name)
                counter += 1
                if counter > 10:
                    raise ValueError('get_path died by hitting recursion limit')
            
        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        
        wb = Workbook()        
        dest_filename = r'table.xlsx'     
        ws = wb.worksheets[0]
        ws.title = "Summary"

        r = 1   
        item = self.tree.GetRootItem()
        while item:
               
            item = self.tree.GetNext(item)
            if not item:
                break
            
            row = get_row(item)
            
            ws.cell('A%s'%(r)).value = '::'.join(get_path(item)+[row[0]])
            for c in range(1,len(row)):
                ws.cell('%s%s' % (get_column_letter(c+1), r)).value = row[c]
            
            r += 1
            
        ws = wb.create_sheet()
        ws.title = 'Pressure profiles'
        
        # Adjust these to adjust the spacing around each block of data
        row_idx = 3
        col_idx = 1
        for run in self.runs:
            
            theta = run.get('summary/theta_profile').value
            p1 = run.get('summary/p1_profile').value
            p2 = run.get('summary/p2_profile').value
            
            # Header
            ws.cell('%s%s' % (get_column_letter(col_idx + 0), row_idx-1)).value = 'theta'
            ws.cell('%s%s' % (get_column_letter(col_idx + 1), row_idx-1)).value = 'path #1'
            ws.cell('%s%s' % (get_column_letter(col_idx + 2), row_idx-1)).value = 'path #2'
                
            # Data
            for r in range(1, len(theta)+1):
                ws.cell('%s%s' % (get_column_letter(col_idx + 0), r + row_idx)).value = theta[r-1]
                ws.cell('%s%s' % (get_column_letter(col_idx + 1), r + row_idx)).value = p1[r-1]
                ws.cell('%s%s' % (get_column_letter(col_idx + 2), r + row_idx)).value = p2[r-1]
        
            col_idx += 3 + 2
        
        for key,name in [('p','Pressure'),('T','Temperature'),('rho','Density'),('V','Volume')]:
            ## Output 
            ws = wb.create_sheet()
            
            ws.title = name
            
            # Adjust these to adjust the spacing around each block of data
            row_idx = 3
            col_idx = 1
            for run in self.runs:
                t = run.get('t').value
                p = run.get(key).value.T
                
                ws.cell('%s%s' % (get_column_letter(col_idx + 0), row_idx-1)).value = 'theta'
                for c in range(1, p.shape[1]+1):
                    CVkey = run.get('CVs/keys/%s' %(c-1,)).value
                    ws.cell('%s%s' % (get_column_letter(col_idx + c), row_idx-1)).value = CVkey
                    
                for r in range(1, len(t)+1):
                    ws.cell('%s%s' % (get_column_letter(col_idx + 0), r + row_idx)).value = t[r-1]
                    for c in range(1, p.shape[1]+1):
                        ws.cell('%s%s' % (get_column_letter(col_idx + c), r + row_idx)).value = p[r-1,c-1]
                        
                col_idx += 3 + p.shape[1]
            
        wb.save(filename = dest_filename)
            
        
    def OnActivate(self, evt):
        
        annotation_dict = self.tree.GetItemPyData(evt.GetItem())
        if annotation_dict:
            self.GetParent().AnnotationTarget.SetLabel('Annotation: ' + annotation_dict['annotation'])
        else:
            self.GetParent().AnnotationTarget.SetLabel('Annotation: None' )
        
        
    def OnRightUp(self, evt):
        pos = evt.GetPosition()
        item, flags, col = self.tree.HitTest(pos)
        #print('Flags: %s, Col:%s, Text: %s' %(flags, col, self.tree.GetItemText(item, col)))

        # Make a menu
        menu = wx.Menu()
        
        # If you are in a data column, allow you to remove the column
        if col > 0:
            menuitem1 = wx.MenuItem(menu, -1, 'Remove this column')
            self.Bind(wx.EVT_MENU, lambda event: self.OnRemoveCol(event, col), menuitem1)
            menu.AppendItem(menuitem1)
            
            menuitem = wx.MenuItem(menu, -1, 'Plot this simulation')
            self.Bind(wx.EVT_MENU, lambda event: self.OnPlotSimulation(event, item, col), menuitem)
            menu.AppendItem(menuitem)
        
        # If 
        if self.tree.GetItemText(item,col).startswith('<HDF5 dataset'):
            _isHDF5array = True
        else:
            _isHDF5array = False
            
        if len(self.tree.GetItemText(item,col)) > 300:
            _isScript = True
        else:
            _isScript = False
        
        if _isHDF5array:
            menuitem = wx.MenuItem(menu, -1, 'Display this parameter')
            self.Bind(wx.EVT_MENU, lambda event: self.OnArrayDisplay(event, item, col), menuitem)
            menu.AppendItem(menuitem)
            menuitem = wx.MenuItem(menu, -1, 'Save to Excel-format csv file')
            self.Bind(wx.EVT_MENU, lambda event: self.OnArray2CSV(event, item, col), menuitem)
            menu.AppendItem(menuitem)
            
        if _isScript:
            menuitem = wx.MenuItem(menu, -1, 'View Script')
            self.Bind(wx.EVT_MENU, lambda event: self.OnScriptDisplay(event, item, col), menuitem)
            menu.AppendItem(menuitem)
             
        # If you are in a dataset row, allow you to sort based on it
        values = [self.tree.GetItemText(item, col) for col in range(1, self.tree.ColumnCount)]
        if any(values):
            menuitem = wx.MenuItem(menu, -1, 'Sort by this parameter')
            self.Bind(wx.EVT_MENU, lambda event: self.OnSortByRow(event, item), menuitem)
            menu.AppendItem(menuitem)
        
        
        if menu.GetMenuItems():
            # Popup the menu.  If an item is selected then its handler
            # will be called before PopupMenu returns.
            self.PopupMenu(menu)
        menu.Destroy()
        
    def OnArray2CSV(self, event, item, col):
        """ Write to a comma separated values file"""
        
        # Build the path to this term
        path = ''
        while not item == self.tree.GetRootItem():
            path = '/' +  self.tree.GetItemText(item, 0) + path
            item = self.tree.GetItemParent(item)
            
        val = self.runs[col-1].get(path).value
        
        FD = wx.FileDialog(None,
                           "Save HDF5 array to file",
                           defaultDir='.',
                           wildcard =  "csv files (*.csv)|*.csv",
                           style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if wx.ID_OK==FD.ShowModal():
            file_path = FD.GetPath()
        else:
            file_path = ''
        FD.Destroy()
        
        # No file selected
        if not file_path:
            return
        
        rows = []
        for i in range(val.shape[0]):
            #  row as a list
            r = val[i,:].tolist()
            #  entries of r as strings
            r = [str(v) for v in r]
            #  add to rows
            rows.append(','.join(r))
        s = '\n'.join(rows)
        
        f = open(file_path, 'w')
        f.write(s)
        f.close()
        
    def OnArrayDisplay(self, event, item, col):
        
        # Title for the viewer
        title = self.tree.GetItemText(item, 0)
        
        # Build the path to this term
        path = ''
        while not item == self.tree.GetRootItem():
            path = '/' +  self.tree.GetItemText(item, 0) + path
            item = self.tree.GetItemParent(item)
            
        val = self.runs[col-1].get(path).value
        
        frm = ArrayDisplay(val, title = title)
        frm.Show()
    
    def OnPlotSimulation(self, event, item, col):
        """
        Plot the simulation
        """
        self.Parent.Parent.plot_outputs(self.runs[col-1])
        self.Parent.Parent.SetSelection(1)
        
    def OnScriptDisplay(self, event, item, col):
        path = ''
        script = self.tree.GetItemText(item, col)
        
        frm = ScriptDisplay(script)
        frm.Show()
        
    def OnSortByRow(self, event, item):
        self.tree.MoveAfterInTabOrder()
        #Skip column 0 which is the header column
        values = [self.tree.GetItemText(item, col) for col in range(1, self.tree.ColumnCount)]
        print 'Not currently implemented'
        print 'row values', values
    
    def OnRemoveCol(self, event, col):
        
        if col > 0:
            self.remove_run(col-1) #Column index 1 is the self.runs index 0
            self.rebuild()

    def OnSize(self, evt):
        self.tree.SetSize(self.GetSize())
        
    def add_runs(self, runs):
        if isinstance(runs, h5py.File):
            self.runs += [runs]
        elif isinstance(runs, basestring) and os.path.exists(runs):
            self.runs += [h5py.File(runs,'r')]
        else:
            for run in runs:
                if isinstance(run, h5py.File):
                    self.runs += [run] 
                elif isinstance(run, basestring) and os.path.exists(run):
                    self.runs += [h5py.File(run,'r')]
                else:
                    raise ValueError("Can only add instances of h5py.File")
        self.rebuild()
        
    def remove_run(self, index):
        """
        Remove the run at the given index
        """
        self.runs.pop(index)
        
class ScriptDisplay(wx.Frame):
    
    def __init__(self, text, **kwargs):
        wx.Frame.__init__(self, None, **kwargs)
        
        panel = wx.Panel(self)
        
        stc = wx.stc.StyledTextCtrl(panel, -1)
    
        import keyword
        stc.SetLexer(wx.stc.STC_LEX_PYTHON)
        stc.SetKeyWords(0, " ".join(keyword.kwlist))
        
        stc.SetText(text)
    
        stc.SetMargins(0,0)
    
        stc.SetViewWhiteSpace(False)
        #self.SetBufferedDraw(False)
        #self.SetViewEOL(True)
        #self.SetEOLMode(stc.STC_EOL_CRLF)
        #self.SetUseAntiAliasing(True)
        
        stc.SetEdgeMode(wx.stc.STC_EDGE_BACKGROUND)
        stc.SetEdgeColumn(78)
    
#        self.Bind(stc.EVT_STC_UPDATEUI, self.OnUpdateUI)
#        self.Bind(stc.EVT_STC_MARGINCLICK, self.OnMarginClick)
#        self.Bind(wx.EVT_KEY_DOWN, self.OnKeyPressed)
    
        faces = {'times': 'Courier New',
              'mono' : 'Courier New',
              'helv' : 'Courier New',
              'other': 'Courier New',
              'size' : 10,
              'size2': 10}
        # Global default styles for all languages
        stc.StyleSetSpec(wx.stc.STC_STYLE_DEFAULT,"face:Courier New,size:10")
        
        # Default 
        stc.StyleSetSpec(wx.stc.STC_P_DEFAULT, "fore:#000000,face:%(helv)s,size:%(size)d" % faces)
        # Comments
        stc.StyleSetSpec(wx.stc.STC_P_COMMENTLINE, "fore:#007F00,face:%(other)s,size:%(size)d" % faces)
        # Number
        stc.StyleSetSpec(wx.stc.STC_P_NUMBER, "fore:#007F7F,size:%(size)d" % faces)
        # String
        stc.StyleSetSpec(wx.stc.STC_P_STRING, "fore:#7F007F,face:%(helv)s,size:%(size)d" % faces)
        # Single quoted string
        stc.StyleSetSpec(wx.stc.STC_P_CHARACTER, "fore:#7F007F,face:%(helv)s,size:%(size)d" % faces)
        # Keyword
        stc.StyleSetSpec(wx.stc.STC_P_WORD, "fore:#00007F,bold,size:%(size)d" % faces)
        # Triple quotes
        stc.StyleSetSpec(wx.stc.STC_P_TRIPLE, "fore:#7F0000,size:%(size)d" % faces)
        # Triple double quotes
        stc.StyleSetSpec(wx.stc.STC_P_TRIPLEDOUBLE, "fore:#7F0000,size:%(size)d" % faces)
        # Class name definition
        stc.StyleSetSpec(wx.stc.STC_P_CLASSNAME, "fore:#0000FF,bold,underline,size:%(size)d" % faces)
        # Function or method name definition
        stc.StyleSetSpec(wx.stc.STC_P_DEFNAME, "fore:#007F7F,bold,size:%(size)d" % faces)
        # Operators
        stc.StyleSetSpec(wx.stc.STC_P_OPERATOR, "bold,size:%(size)d" % faces)
        # Identifiers
        stc.StyleSetSpec(wx.stc.STC_P_IDENTIFIER, "fore:#000000,face:%(helv)s,size:%(size)d" % faces)
        # Comment-blocks
        stc.StyleSetSpec(wx.stc.STC_P_COMMENTBLOCK, "fore:#7F7F7F,size:%(size)d" % faces)
        # End of line where string is not closed
        stc.StyleSetSpec(wx.stc.STC_P_STRINGEOL, "fore:#000000,face:%(mono)s,back:#E0C0E0,eol,size:%(size)d" % faces)

        stc.SetCaretForeground("BLUE")
        
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(stc,1,wx.EXPAND)
        panel.SetSizer(sizer)
        
        # Frame layout
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        main_sizer.Add(panel,1, wx.EXPAND)
        self.SetSizer(main_sizer)
        main_sizer.Layout()
    
class ArrayDisplay(wx.Frame):
    
    def __init__(self, array, **kwargs):
        
        wx.Frame.__init__(self, None, **kwargs)
        
        panel = wx.Panel(self)
        
        array = np.array(array, ndmin = 2)
        self.grid = wx.grid.Grid(panel)

        # Make the grid the same shape as the data
        self.grid.CreateGrid(*array.shape)
        
        # Values
        for row in range(array.shape[0]):
            for col in range(array.shape[1]):
                self.grid.SetCellValue(row,col,str(array[row][col]))
                
        # Column headers
        for col in range(array.shape[1]):
            self.grid.SetColLabelValue(col,str(col))
        
        # Row "headers"
        for row in range(array.shape[0]):
            self.grid.SetRowLabelValue(row,str(row))
                
        # Panel layout
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.grid,1,wx.EXPAND)
        panel.SetSizer(sizer)
        
        # Frame layout
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        main_sizer.Add(panel,1, wx.EXPAND)
        self.SetSizer(main_sizer)
        main_sizer.Layout()
        
class ParaSelectDialog(wx.Dialog):
    def __init__(self):
        wx.Dialog.__init__(self, None, title = "State Chooser",)
        self.MinLabel, self.Min = LabeledItem(self, label = 'Minimum value', value = '')
        self.MaxLabel, self.Max = LabeledItem(self, label = 'Maximum value', value = '')
        self.StepsLabel, self.Steps = LabeledItem(self, label = 'Number of steps', value = '')
        self.Accept = wx.Button(self, label = "Accept")
        sizer = wx.FlexGridSizer(cols = 2, hgap = 4, vgap = 4)
        sizer.AddMany([self.MinLabel, self.Min, self.MaxLabel, 
                       self.Max, self.StepsLabel, self.Steps, self.Accept])
        self.SetSizer(sizer)
        sizer.Layout()
        self.Fit()
        self.Accept.Bind(wx.EVT_BUTTON, self.OnAccept)
        
        #Bind a key-press event to all objects to get Esc 
        children = self.GetChildren()
        for child in children:
            child.Bind(wx.EVT_KEY_UP,  self.OnKeyPress)
        
    def join_values(self):
        values = np.linspace(float(self.Min.GetValue()),float(self.Max.GetValue()),int(self.Steps.GetValue()))
        return ', '.join([str(val) for val in values]) 
        
    def OnAccept(self, event = None):
        self.EndModal(wx.ID_OK)
        
    def OnKeyPress(self,event = None):
        """ cancel if Escape key is pressed """
        event.Skip()
        if event.GetKeyCode() == wx.WXK_ESCAPE:
            self.EndModal(wx.ID_CANCEL)
    
    def CancelValues(self, event = None):
        self.EndModal(wx.ID_CANCEL)


class ParametricOption(wx.Panel):
    
    
    def __init__(self, parent, GUI_objects):
        wx.Panel.__init__(self, parent)
        
        labels = [o.annotation for o in GUI_objects.itervalues()]
        
        # Check that there is no duplication between annotations
        if not len(labels) == len(set(labels)): # Sets do not allow duplication
            raise ValueError('You have duplicated annotations which is not allowed')
        
        # Make a reverse map from annotation to GUI object key
        self.GUI_map = {o.annotation:o.key for o in GUI_objects.itervalues()} 
        
        self.Terms = wx.ComboBox(self)
        self.Terms.AppendItems(sorted(labels))
        self.Terms.SetSelection(0)
        self.Terms.SetEditable(False)
        self.RemoveButton = wx.Button(self, label = '-', style = wx.ID_REMOVE)
        
        self.Values = wx.TextCtrl(self, value = '')
        self.Select = wx.Button(self, label = 'Select...')
        
        sizer = wx.BoxSizer(wx.HORIZONTAL)
        sizer.Add(self.RemoveButton)
        sizer.Add(self.Terms)
        sizer.Add(self.Values)
        sizer.Add(self.Select)
        self.SetSizer(sizer)
        
        self.Select.Bind(wx.EVT_BUTTON,self.OnSelectValues)
        self.RemoveButton.Bind(wx.EVT_BUTTON, lambda event: self.Parent.RemoveTerm(self))
        self.Terms.Bind(wx.EVT_COMBOBOX, self.OnChangeTerm)
        
        self.OnChangeTerm()
        
    def OnChangeTerm(self, event = None):
        """
        Update the values when the term is changed if a structured table
        """
        if self.GetParent().Structured.GetValue():
            annotation = self.Terms.GetStringSelection()
            
            # Get the key of the registered object
            key = self.GUI_map[annotation]
            
            # Get the value of the item in the GUI
            val = self.GetTopLevelParent().get_GUI_object_value(key)
            
            # Set the textctrl with this value
            self.Values.SetValue(str(val))
    
    def OnSelectValues(self, event = None):
        dlg = ParaSelectDialog()
        if dlg.ShowModal() == wx.ID_OK:
            self.Values.SetValue(dlg.join_values())
        dlg.Destroy()
        
    def get_values(self):
        """
        Get a list of floats from the items in the textbox
        
        Returns
        -------
        annotation
        vals
        """
        annotation = self.Terms.GetStringSelection()
        #To list of floats
        if hasattr(self,'Values'):
            values = [float(val) for val in self.Values.GetValue().split(',') if not val == '']
        else:
            values = None
        return annotation, values
    
    def set_values(self,key,vals):
        """
        Parameters
        ----------
        key : the registered term key
        vals : A list of values as floats
        """
        annotation = self.GetTopLevelParent().get_GUI_object(key).annotation
        self.Terms.SetStringSelection(annotation)
        if hasattr(self,'Values'):
            strvalues = ', '.join([str(v) for v in vals])
            self.Values.SetValue(strvalues)
        
    def make_unstructured(self):
        if hasattr(self,'Values'):
            self.Values.Destroy()
            del self.Values
        if hasattr(self,'Select'):
            self.Select.Destroy()
            del self.Select
            
        self.GetSizer().Layout()
        self.Refresh()
    
    def make_structured(self):
        if not hasattr(self,'Values'):
            self.Values = wx.TextCtrl(self, value = '')
            self.Select = wx.Button(self, label = 'Select...')
            self.Select.Bind(wx.EVT_BUTTON,self.OnSelectValues)
            self.GetSizer().AddMany([self.Values,self.Select])
            self.GetSizer().Layout()
            self.Refresh()
            self.OnChangeTerm()
        
class ParametricCheckList(wx.ListCtrl, ListCtrlAutoWidthMixin, CheckListCtrlMixin, TextEditMixin):
    """
    The checklist that stores all the possible runs
    """
    def __init__(self, parent, headers, values, structured = True):
        """
        Parameters
        ----------
        parent : wx.window
            The parent of this checklist
        headers : list
            A list of header strings
        values : 
        structured : bool
            If ``True``, use a structured parametric table to do the calcs
        """
        wx.ListCtrl.__init__(self, parent, -1, style=wx.LC_REPORT)
        ListCtrlAutoWidthMixin.__init__(self)
        CheckListCtrlMixin.__init__(self)
        TextEditMixin.__init__(self)
        
        
        #Build the headers
        self.InsertColumn(0, '')
        self.SetColumnWidth(0, wx.LIST_AUTOSIZE)
        for i, header in enumerate(headers):
            self.InsertColumn(i+1, header)
        self.headers = headers
        
        if not structured:
            #Transpose the nested lists
            self.data = zip(*values)
            #Turn back from tuples to lists
            self.data = [list(row) for row in self.data]
        else:
            self.data = [list(row) for row in itertools.product(*values)]
        
        #Add the values one row at a time
        for i,row in enumerate(self.data):
            self.InsertStringItem(i,'')
            for j,val in enumerate(row):
                self.SetStringItem(i,j+1,str(val))
            self.CheckItem(i)
            
        for i in range(len(headers)):
            self.SetColumnWidth(i+1,wx.LIST_AUTOSIZE_USEHEADER)
        self.Bind(wx.EVT_LIST_ITEM_ACTIVATED, self.OnItemActivated)
        self.Bind(wx.EVT_LIST_BEGIN_LABEL_EDIT, self.PreCellEdit)
        self.Bind(wx.EVT_LIST_END_LABEL_EDIT, self.OnCellEdited)

    def OnItemActivated(self, event):
        self.ToggleItem(event.m_itemIndex)
    
    def PreCellEdit(self, event):
        """
        Before the cell is edited, only allow edits on columns after the first one
        """
        row_index = event.m_itemIndex
        col_index = event.Column
        if col_index == 0:
            event.Veto()
        else:
            val = float(event.Text)
            self.data[row_index][col_index-1] = val
            event.Skip()
    
    def OnCellEdited(self, event):
        """
        Once the cell is edited, write its value back into the data matrix
        """
        row_index = event.m_itemIndex
        col_index = event.Column
        val = float(event.Text)
        self.data[row_index][col_index-1] = val
        
    def SetCellValue(self,Irow,Icol,val):
        """ Set the string value of the cell at Irow,Icol """ 
        self.data[Irow][Icol] = val
        self.SetStringItem(Irow,Icol+1,val)
    
    def GetStringCell(self,Irow,Icol):
        """ Returns a string representation of the cell """
        return self.data[Irow][Icol]
    
    def GetFloatCell(self,Irow,Icol):
        """ Returns a float representation of the cell """
        return float(self.data[Irow][Icol])
    
    def AddRow(self):
        
        row = [0]*(self.GetColumnCount()-1)
        
        i = len(self.data)
        self.InsertStringItem(i,'')
        for j,val in enumerate(row):
            self.SetStringItem(i,j+1,str(val))
        self.CheckItem(i)
        
        self.data.append(row)
        
    def RemoveRow(self, i = 0):
        self.data.pop(i)
        self.DeleteItem(i)
        
    def RemoveLastRow(self):
        i = len(self.data)-1
        self.data.pop(i)
        self.DeleteItem(i)
                               
class HackedButton(wx.Button):
    """
    This is needed because of a bug in FlatMenu where clicking on a disabled 
    item throws an exception
    """
    def __init__(self, parent, *args, **kwargs):
        wx.Button.__init__(self, parent, *args, **kwargs)
        
        self.Enable()
        
    def Enable(self):
        self.SetForegroundColour((70,70,70))
        self._Enabled = True
        self.SetEvtHandlerEnabled(True)
        
    def Disable(self):
        self.SetForegroundColour((255,255,255))
        self._Enabled = False
        self.SetEvtHandlerEnabled(False)
         
class ParametricPanel(PDPanel):
    """
    The top-level panel to contain all the things for the parametric table
    """
    def __init__(self, parent, configdict, **kwargs):
        PDPanel.__init__(self, parent, **kwargs)
        
        import wx.lib.agw.flatmenu as FM
        from wx.lib.agw.fmresources import FM_OPT_SHOW_CUSTOMIZE, FM_OPT_SHOW_TOOLBAR, FM_OPT_MINIBAR
        self._mb = FM.FlatMenuBar(self, wx.ID_ANY, 10, 5, options = FM_OPT_SHOW_TOOLBAR)
        
        self.Structured = wx.CheckBox(self._mb, label = 'Structured')
        self._mb.AddControl(self.Structured)
        self.Structured.Bind(wx.EVT_CHECKBOX, self.OnChangeStructured)
        self.Structured.SetValue(configdict['structured'])
        
        self.AddButton = HackedButton(self._mb, label = 'Add Term')
        self._mb.AddControl(self.AddButton)
        self.AddButton.Bind(wx.EVT_BUTTON, self.OnAddTerm)
        
        self.BuildButton = HackedButton(self._mb, label = 'Build Table')
        self._mb.AddControl(self.BuildButton)
        self.BuildButton.Bind(wx.EVT_BUTTON, self.OnBuildTable)
        self.BuildButton.Disable()
        
        self.RunButton = HackedButton(self._mb, label = 'Run Table')
        self._mb.AddControl(self.RunButton)
        self.RunButton.Bind(wx.EVT_BUTTON, self.OnRunTable)
        self.RunButton.Disable()
        
        self.ZipButton = HackedButton(self._mb, label = 'Make Batch .zip')
        self._mb.AddControl(self.ZipButton)
        self.ZipButton.Bind(wx.EVT_BUTTON, self.OnZipBatch)
        self.ZipButton.Disable()
        
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self._mb,0,wx.EXPAND)
        self.SetSizer(sizer)
        sizer.Layout()
        
        self.NTerms = 0
        self.ParamSizer = None
        self.ParamListSizer = None
        self.ParaList = None
        
        self.GUI_map = {o.annotation:o.key for o in self.main.get_GUI_object_dict().itervalues()} 
        
        # Populate the terms from the configuration dictionary 
        self.populate_terms(configdict)
        
        # After all the building is done, check if it is unstructured, if so, 
        # collect the temporary values that were set 
#        if self.Structured.GetValue() == True:
#            self.OnBuildTable()
        self.Layout()
        self.OnChangeStructured()
            
    def OnChangeStructured(self, event = None):
        
        #Param
        terms = [term for term in self.Children if isinstance(term,ParametricOption)]
        
        for term in terms:
            if self.Structured.GetValue() == True:
                term.make_structured()
            else:
                term.make_unstructured()

        self.GetSizer().Layout()
        self.Refresh()
            
    def OnAddTerm(self, event = None):
        if self.NTerms == 0:
            self.ParamSizerBox = wx.StaticBox(self, label = "Parametric Terms")
            self.ParamSizer = wx.StaticBoxSizer(self.ParamSizerBox, wx.VERTICAL)
            self.GetSizer().Add(self.ParamSizer)
            self.BuildButton.Enable()
        # Get all the registered GUI objects
        GUI_objects = self.main.get_GUI_object_dict()
        # Create the option panel
        option = ParametricOption(self, GUI_objects)
        # Make it either structured or unstructured
        if self.Structured.GetValue() == True:
            option.make_structured()
        else:
            option.make_unstructured()
        
        self.ParamSizer.Add(option)
        self.ParamSizer.Layout()
        self.NTerms += 1
        self.GetSizer().Layout()
        self.Refresh()
        return option
    
    def RemoveTerm(self, term):
        term.Destroy()
        self.NTerms -= 1
        if self.NTerms == 0:
            if self.ParamSizer is not None:
                self.GetSizer().Remove(self.ParamSizer)
                self.ParamSizer = None
            if self.ParaList is not None:
                self.ParaList.Destroy()
                self.ParaList = None
            if self.ParamListSizer is not None:
                self.GetSizer().Remove(self.ParamListSizer)
                self.ParamListSizer = None
            self.RunButton.Disable()
            self.BuildButton.Disable()
            self.ZipButton.Disable()
        else:
            self.ParamSizer.Layout()
        self.GetSizer().Layout()
        self.Refresh()
        
    def OnBuildTable(self, event=None):
        names = []
        values = []
        #make names a list of strings
        #make values a list of lists of values
        for param in self.ParamSizer.GetChildren():
            name, vals = param.Window.get_values()
            names.append(name)
            if self.Structured.GetValue() == True:
                values.append(vals)
            else:
                if hasattr(param.Window,'temporary_values'):
                    values.append(param.Window.temporary_values.split(';'))
                else:
                    values.append(['0.0'])
        
        #Build the list of parameters for the parametric study
        if self.ParamListSizer is None:
            self.ParamListBox = wx.StaticBox(self, label = "Parametric Terms Ranges")
            self.ParamListSizer = wx.StaticBoxSizer(self.ParamListBox, wx.VERTICAL)
        else:
            self.ParaList.Destroy()
            self.GetSizer().Remove(self.ParamListSizer)
#            #Build and add a sizer for the para values
            self.ParamListBox = wx.StaticBox(self, label = "Parametric Runs")
            self.ParamListSizer = wx.StaticBoxSizer(self.ParamListBox, wx.VERTICAL)
        
        #Remove the spinner and its friends
        if hasattr(self,'RowCountSpinner'):
            self.RowCountSpinner.Destroy(); del self.RowCountSpinner
            self.RowCountLabel.Destroy(); del self.RowCountLabel
            self.RowCountSpinnerText.Destroy(); del self.RowCountSpinnerText
            
        #Build and add a sizer for the para values
        if self.Structured.GetValue() == False:
            self.RowCountLabel = wx.StaticText(self,label='Number of rows')
            self.RowCountSpinnerText = wx.TextCtrl(self, value = "1", size = (40,-1))
            h = self.RowCountSpinnerText.GetSize().height
            w = self.RowCountSpinnerText.GetSize().width + self.RowCountSpinnerText.GetPosition().x + 2
            self.RowCountSpinner = wx.SpinButton(self, wx.ID_ANY, (w,50), (h*2/3,h), style = wx.SP_VERTICAL)
            self.RowCountSpinner.SetRange(1, 1000)
            self.RowCountSpinner.SetValue(1)
            self.RowCountSpinner.Bind(wx.EVT_SPIN_UP, self.OnSpinUp)
            self.RowCountSpinner.Bind(wx.EVT_SPIN_DOWN, self.OnSpinDown)
            
            sizer = wx.BoxSizer(wx.HORIZONTAL)
            sizer.AddMany([self.RowCountLabel, self.RowCountSpinner,self.RowCountSpinnerText])
            self.GetSizer().Add(sizer)
                    
        self.GetSizer().Add(self.ParamListSizer,1,wx.EXPAND)
        self.ParaList = ParametricCheckList(self,names,values,
                                            structured = self.Structured.GetValue())
            
        self.ParamListSizer.Add(self.ParaList,1,wx.EXPAND)
        self.ParaList.SetMinSize((400,-1))
        self.ParamListSizer.Layout()
        self.GetSizer().Layout()
        self.Refresh() 
        
        # Enable the batch buttons
        self.RunButton.Enable()
        self.ZipButton.Enable()
            
        if self.Structured.GetValue() == False:
            self.RowCountSpinner.SetValue(self.ParaList.GetItemCount())
            self.RowCountSpinnerText.SetValue(str(self.ParaList.GetItemCount()))
            #Bind a right click to opening a popup
            # for wxMSW
            self.ParaList.Bind(wx.EVT_COMMAND_RIGHT_CLICK, self.OnParaListRightClick)
            # for wxGTK
            self.ParaList.Bind(wx.EVT_RIGHT_UP, self.OnParaListRightClick)
            
    def OnParaListRightClick(self, event): 
        #based on wxpython demo program
        
        # make a menu
        menu = wx.Menu()
        # add some items
        menuitem1 = wx.MenuItem(menu, -1, 'Fill from clipboard')
        menuitem2 = wx.MenuItem(menu, -1, 'Fill from csv file (future)')
        
        self.Bind(wx.EVT_MENU, self.OnPaste, menuitem1)
        menu.AppendItem(menuitem1)
        menu.AppendItem(menuitem2)
        
        # Popup the menu.  If an item is selected then its handler
        # will be called before PopupMenu returns.
        self.PopupMenu(menu)
        menu.Destroy()
        
    def OnPaste(self, event):
        """
        Paste the contents of the clipboard into the table
        """
        do = wx.TextDataObject()
        if wx.TheClipboard.Open():
            success = wx.TheClipboard.GetData(do)
            wx.TheClipboard.Close()

        data = do.GetText()
        rows = data.strip().replace('\r','').split('\n')
        rows = [row.split('\t') for row in rows]
        #Check that the dimensions of pasted section and table are the same
        if not self.ParaList.GetItemCount() == len(rows):
            msg = 'There are '+str(len(rows))+' rows in your pasted table, but '+str(self.ParaList.GetItemCount())+' rows in the table.  Resize table to fit?'
            dlg = wx.MessageDialog(None, msg)
            if dlg.ShowModal() == wx.ID_OK:
                #  Resize table to fit
                if len(rows) < self.ParaList.GetItemCount():
                    while self.ParaList.GetItemCount() > len(rows):
                        self.OnSpinDown()
                if len(rows) > self.ParaList.GetItemCount():
                    while self.ParaList.GetItemCount() < len(rows):
                        self.OnSpinUp()
                dlg.Destroy()
            else:
                dlg.Destroy()
                return
            
        #Right number of rows, set the values
        for i,row in enumerate(rows): 
            for j,item in enumerate(row): 
                self.ParaList.SetStringItem(i,j+1,item)
                self.ParaList.data[i][j] = item
        
    def OnSpinDown(self, event = None):
        """ 
        Fires when the spinner is used to decrease the number of rows
        """
        
        #  Remove the last row
        self.ParaList.RemoveLastRow()
            
        #  Set the textbox value
        self.RowCountSpinner.SetValue(self.ParaList.GetItemCount())
        self.RowCountSpinnerText.SetValue(str(self.ParaList.GetItemCount()))
        
        
    def OnSpinUp(self, event = None):
        """ 
        Fires when the spinner is used to increase the number of rows
        """
        
        #  Add a row
        self.ParaList.AddRow()
        
        #  Set the textbox value
        self.RowCountSpinner.SetValue(self.ParaList.GetItemCount())
        self.RowCountSpinnerText.SetValue(str(self.ParaList.GetItemCount()))
        
    def build_all_scripts(self):
        sims = []
        
        # Column index 1 is the list of parameters
        for Irow in range(self.ParaList.GetItemCount()):
            # Loop over all the rows that are checked
            if self.ParaList.IsChecked(Irow):
                
                #Empty lists for this run
                vals, names = [], []
                
                # Each row corresponds to one run of the model
                # 
                # Loop over the columns for this row 
                for Icol in range(self.ParaList.GetColumnCount()-1):
                    vals.append(self.ParaList.GetFloatCell(Irow, Icol))
                    names.append(self.ParaList.GetColumn(Icol+1).Text)
                    
                # The attributes corresponding to the names
                keys = [self.GUI_map[name] for name in names]
                
                # Get a list of the objects that are coupled with other GUI objects
                coupled_objects = []
                for k in keys:
                    if isinstance(self.main.get_GUI_object(k), CoupledAnnotatedGUIObject):
                        coupled_objects.append(self.main.get_GUI_object(k))
                            
                # First check if any of the terms are coupled gui objects
                if coupled_objects:
                    # If they are, handle them differently
                    for o in coupled_objects:
                        # Check the partners exist
                        for o in o.required_partners:
                            if o not in coupled_objects:
                                raise KeyError("Coupled object [{o:s}] not found".format(o = o))
                    
                    #Pass all the coupled objects to the handler
                    o.handler(coupled_objects,keys,vals)
                else:
                    print 'no coupled objects'
                
                # Set the value in the GUI
                for key,val in zip(keys,vals):
                    self.main.set_GUI_object_value(key, val)
                
                #Build the simulation script using the GUI parameters
                script_name = self.main.build_simulation_script(run_index = Irow)
                            
                #Add sim to the list (actually append the path to the script)
                sims.append(script_name)
                
        # Check that there is a difference between the files generated 
        if not self.check_scripts_are_different(sims):
            dlg = wx.MessageDialog(None,'Cannot run batch because some of the batch files are exactly the same. Deleting generated scripts')
            dlg.ShowModal()
            dlg.Destroy()
            
            for file in sims:
                # Full path to file
                fName = os.path.join(PDSimGUI.pdsim_home_folder,file)
                # Remove the file generated, don't do anything if error
                try:
                    os.unlink(fName)
                except OSError:
                    pass
            
            return []
        
        return sims
        
    def OnRunTable(self, event=None):
        """
        Actually runs the parametric table
        
        This event can only fire if the table is built
        """
        
        #Build all the scripts
        sims = self.build_all_scripts()
        
        if sims:
            #Actually run the batch with the sims that have been built
            self.main.run_batch(sims)
        
    def check_scripts_are_different(self, scripts):
        """
        return ``True`` if the scripts differ by more than the time stamp, ``False`` otherwise
        
        The first 10 lines are not considered in the diff
        """
        
        # Take the lines that follow the first 10 lines and read into memory
        chopped_scripts = [open(os.path.join(PDSimGUI.pdsim_home_folder,fName),'r').readlines()[10::] for fName in scripts]
        
        # Compare each file with all other files
        for i in range(len(chopped_scripts)):
            for j in range(i+1,len(chopped_scripts)):
                # If the list of unified diffs is empty, the files are the same
                # This is a failure
                if not [d for d in difflib.unified_diff(chopped_scripts[i],chopped_scripts[j])]:
                    return False
        # Made it this far, return True, all the files are different
        return True
    
    def OnZipBatch(self, event = None):
        """
        Write all the script files and a runner to a zip file - this can be
        nice to do batches outside the GUI.
        """
        
        template = textwrap.dedent(
        """
        import glob, os
        from PDSim.misc.hdf5 import HDF5Writer
        
        H = HDF5Writer()
        
        for file in glob.glob('script_*.py'):
            root,ext = os.path.splitext(file)
            mod = __import__(root)
            sim = mod.build()
            mod.run(sim)
            
            #  Remove FlowsStorage as it is enormous
            del sim.FlowStorage
            
            H.write_to_file(sim,root+'.h5')
               
           """
           )
        
        sims = self.build_all_scripts()
        
        if sims:
            
            FD = wx.FileDialog(None,
                               "Save zip file",
                               defaultDir='.',
                               wildcard =  "ZIP files (*.zip)|*.zip|All Files|*.*",
                               style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
            if wx.ID_OK==FD.ShowModal():
                zip_file_path = FD.GetPath()
            else:
                zip_file_path = ''
            FD.Destroy()
        
            if zip_file_path:
                with zipfile.ZipFile(zip_file_path,'w') as z:
                    for file in sims:
                        # Full path to file
                        fName = os.path.join(PDSimGUI.pdsim_home_folder,file)
                        # write the file, strip off the path
                        z.write(fName, arcname = file)
                    
                    z.writestr('run.py',template)
        
    def populate_terms(self,configdict):
        """
        Load the terms if there are any
        """
        if 'terms' not in configdict:
            return
        
        for term in configdict['terms']:
            option = self.OnAddTerm()
            option.set_values(term['key'], term['vals'])
            
        if (self.BuildButton._Enabled):
            self.OnBuildTable()
            
        if not configdict['structured']:
            # Add the necessary rows
            if len(configdict['terms']) > 0:
                for i,v in enumerate(configdict['terms'][0]['vals']):
                    if i > 0:
                        self.ParaList.AddRow()
                        self.NTerms += 1
            
            for i,term in enumerate(configdict['terms']):
                for j,v in enumerate(term['vals']):
                    self.ParaList.SetCellValue(j,i,str(v))
                    
            self.RowCountSpinner.SetValue(self.ParaList.GetItemCount())
            self.RowCountSpinnerText.SetValue(str(self.ParaList.GetItemCount()))
        
    def get_config_chunk(self):
        
        configdict = {}
        configdict['structured'] = self.Structured.GetValue()
        
        if not hasattr(self, 'ParaList'):
            return configdict
        
        terms = []
        if self.Structured.GetValue():
            for i, child in enumerate(self.GetChildren()):
                if isinstance(child, ParametricOption):
                    annotation, vals = child.get_values()
                    key = self.GUI_map[annotation]
                    terms.append(dict(key = key, vals = vals))
            configdict['terms'] = terms
        else:
            for i, head in enumerate(self.ParaList.headers):
                key = self.GUI_map[head]
                vals = [self.ParaList.data[j][i] for j in range(len(self.ParaList.data))]
                terms.append(dict(key = key, vals = vals))
            configdict['terms'] = terms
        
        return configdict

def LabeledItem(parent,id=-1, label='A label', value='0.0', enabled=True, tooltip = None):
    """
    A convenience function that returns a tuple of StaticText and TextCtrl 
    items with the necessary label and values set
    """
    label = wx.StaticText(parent,id,label)
    thing = wx.TextCtrl(parent,id,value)
    if enabled==False:
        thing.Disable()
    if tooltip is not None:
        if enabled:
            thing.SetToolTipString(tooltip)
        else:
            label.SetToolTipString(tooltip)
    return label,thing

class StateChooser(wx.Dialog):
    """
    A dialog used to select the state
    """
    def __init__(self,Fluid,T,rho,parent=None,id=-1,Fluid_fixed = False):
        wx.Dialog.__init__(self,parent,id,"State Chooser",size=(300,250))
        
        class StateChoices(wx.Choicebook):
            def __init__(self, parent, id=-1,):
                wx.Choicebook.__init__(self, parent, id)
                
                self.pageT_dTsh=wx.Panel(self)
                self.AddPage(self.pageT_dTsh,'Saturation Temperature and Superheat')
                self.Tsatlabel1, self.Tsat1 = LabeledItem(self.pageT_dTsh,label="Saturation Temperature [K]",value='290')
                self.Tsat1.default_units = 'Kelvin'
                self.Tsat1.Bind(wx.EVT_CONTEXT_MENU,self.OnChangeUnits)
                self.DTshlabel1, self.DTsh1 = LabeledItem(self.pageT_dTsh,label="Superheat [K]",value='11.1')
                sizer=wx.FlexGridSizer(cols=2,hgap=3,vgap=3)
                sizer.AddMany([self.Tsatlabel1, self.Tsat1,self.DTshlabel1, self.DTsh1])
                self.pageT_dTsh.SetSizer(sizer)
                
                self.pageT_p=wx.Panel(self)
                self.AddPage(self.pageT_p,'Temperature and Absolute Pressure')
                self.Tlabel1, self.T1 = LabeledItem(self.pageT_p,label="Temperature [K]",value='300')
                self.T1.default_units = 'Kelvin'
                self.T1.Bind(wx.EVT_CONTEXT_MENU,self.OnChangeUnits)
                self.plabel1, self.p1 = LabeledItem(self.pageT_p,label="Pressure [kPa]",value='300')
                self.p1.default_units = 'kPa'
                self.p1.Bind(wx.EVT_CONTEXT_MENU,self.OnChangeUnits)
                sizer=wx.FlexGridSizer(cols=2,hgap=3,vgap=3)
                sizer.AddMany([self.Tlabel1, self.T1,self.plabel1, self.p1])
                self.pageT_p.SetSizer(sizer)
            
            def OnChangeUnits(self, event):
                TextCtrl = event.GetEventObject()
                dlg = UnitConvertor(value = float(TextCtrl.GetValue()),
                                    default_units = TextCtrl.default_units
                                    )
                dlg.ShowModal()
                TextCtrl.SetValue(dlg.get_value())
                dlg.Destroy()
        
        sizer=wx.BoxSizer(wx.VERTICAL)
        self.Fluidslabel = wx.StaticText(self,-1,'Fluid: ')
        self.Fluids = wx.ComboBox(self,-1)
        self.Fluids.AppendItems(sorted(CoolProp.__fluids__))
        self.Fluids.SetEditable(False)
        self.Fluids.SetValue(Fluid)
        if Fluid_fixed:
            self.Fluids.Enable(False)
        
        hs = wx.BoxSizer(wx.HORIZONTAL)
        hs.AddMany([self.Fluidslabel,self.Fluids])
        sizer.Add(hs)
        
        sizer.Add((5,5))
        
        self.SC=StateChoices(self)
        sizer.Add(self.SC,1,wx.EXPAND)                
        
        fgs= wx.FlexGridSizer(cols=2,hgap=3,vgap=3)
        self.Tlabel, self.T = LabeledItem(self,label="Temperature [K]",value='300',enabled=False)
        self.plabel, self.p = LabeledItem(self,label="Pressure [kPa]",value='300',enabled=False)
        self.rholabel, self.rho = LabeledItem(self,label="Density [kg/m³]",value='1',enabled=False)
        fgs.AddMany([self.Tlabel,self.T,self.plabel,self.p,self.rholabel,self.rho])
        sizer.Add(fgs)
        
        self.cmdAccept = wx.Button(self,-1,"Accept")
        sizer.Add(self.cmdAccept)
        
        self.SetSizer(sizer)
        self.Fluids.SetStringSelection(Fluid)
        
        if CP.Props(Fluid,"Ttriple") < T < CP.Props(Fluid,"Tcrit"):
            #Pressure from temperature and density
            p = CP.Props('P','T',T,'D',rho,Fluid)
            #Saturation temperature
            Tsat = CP.Props('T','P',p,'Q',1,Fluid)
            self.SC.Tsat1.SetValue(str(Tsat))
            self.SC.DTsh1.SetValue(str(T-Tsat))
            self.SC.T1.SetValue(str(T))
            self.SC.p1.SetValue(str(p))
            self.SC.SetSelection(0) ## The page of Tsat,DTsh
        else:
            #Pressure from temperature and density
            p = CP.Props('P','T',T,'D',rho,Fluid)
            self.SC.T1.SetValue(str(T))
            self.SC.p1.SetValue(str(p))
            self.SC.SetSelection(1) ## The page of Tsat,DTsh
        
        self.OnUpdateVals()
        
        self.SC.Tsat1.Bind(wx.EVT_KEY_UP,self.OnUpdateVals)
        self.SC.DTsh1.Bind(wx.EVT_KEY_UP,self.OnUpdateVals)
        self.SC.T1.Bind(wx.EVT_KEY_UP,self.OnUpdateVals)
        self.SC.p1.Bind(wx.EVT_KEY_UP,self.OnUpdateVals)
        
        self.Fluids.Bind(wx.EVT_COMBOBOX, self.OnFlushVals)
        self.Bind(wx.EVT_CLOSE,self.CancelValues)
        self.cmdAccept.Bind(wx.EVT_BUTTON,self.AcceptValues)
        
        #Bind a key-press event to all objects to get Esc 
        children = self.GetChildren()
        for child in children:
            child.Bind(wx.EVT_KEY_UP,  self.OnKeyPress)
        
    def OnFlushVals(self,event=None):
        """ Clear all the values"""
        self.SC.Tsat1.SetValue("")
        self.SC.DTsh1.SetValue("")
        self.SC.T1.SetValue("")
        self.SC.p1.SetValue("")
        self.T.SetValue("")
        self.p.SetValue("")
        self.rho.SetValue("")
        
    def OnKeyPress(self,event=None):
        """ cancel if Escape key is pressed """
        event.Skip()
        if event.GetKeyCode() == wx.WXK_ESCAPE:
            self.EndModal(wx.ID_CANCEL)
    
    def CancelValues(self,event=None):
        self.EndModal(wx.ID_CANCEL)
        
    def AcceptValues(self,event=None):
        """ If the state is in the vapor phase somewhere, accept it and return """
        Fluid = str(self.Fluids.GetStringSelection())
        T=float(self.T.GetValue())
        p=float(self.p.GetValue())
        if CP.Phase(Fluid,T,p) not in ['Gas','Supercritical']:
            dlg = wx.MessageDialog(None, message = "The phase is not gas or supercritical, cannot accept this state",caption='Invalid state')
            dlg.ShowModal()
            dlg.Destroy()
            return
        self.EndModal(wx.ID_OK)
    
    def GetValues(self):
        Fluid=str(self.Fluids.GetStringSelection())
        T=float(self.T.GetValue())
        p=float(self.p.GetValue())
        rho=float(self.rho.GetValue())
        return Fluid,T,p,rho
        
    def OnUpdateVals(self,event=None):
        if event is not None:
            event.Skip()
            
        PageNum = self.SC.GetSelection()
        Fluid = str(self.Fluids.GetStringSelection())
        try:
            if PageNum == 0:
                #Sat temperature and superheat are given
                p=CP.Props('P','T',float(self.SC.Tsat1.GetValue()),'Q',1.0,Fluid)
                T=float(self.SC.Tsat1.GetValue())+float(self.SC.DTsh1.GetValue())
                rho=CP.Props('D','T',T,'P',p,Fluid)
            elif PageNum == 1:
                #Temperature and pressure are given
                T=float(self.SC.T1.GetValue())
                p=float(self.SC.p1.GetValue())
                rho=CP.Props('D','T',T,'P',p,Fluid)
            else:
                raise NotImplementedError
            
            self.T.SetValue(str(T))
            self.p.SetValue(str(p))
            self.rho.SetValue(str(rho))
        except ValueError:
            return
        
class StatePanel(wx.Panel):
    """
    This is a generic Panel that has the ability to select a state given by 
    Fluid, temperature and density by selecting the desired set of inputs in a
    dialog which can be Tsat and DTsh or T & p.
    """
    def __init__(self, parent, id = -1, CPState = None, Fluid_fixed = False):
        wx.Panel.__init__(self, parent, id)
        
        # If the fluid is not allowed to be changed Fluid_fixed is true
        self._Fluid_fixed = Fluid_fixed
        
        sizer = wx.FlexGridSizer(cols=2,hgap=4,vgap=4)
        self.Fluidlabel, self.Fluid = LabeledItem(self,label="Fluid",value=str(CPState.Fluid))
        
        self.Tlabel, self.T = LabeledItem(self,label="Temperature [K]",value=str(CPState.T))
        self.plabel, self.p = LabeledItem(self,label="Pressure [kPa]",value=str(CPState.p))
        self.rholabel, self.rho = LabeledItem(self,label="Density [kg/m³]",value=str(CPState.rho))
        sizer.AddMany([self.Fluidlabel, self.Fluid,
                       self.Tlabel, self.T,
                       self.plabel, self.p,
                       self.rholabel, self.rho])
        
        for box in [self.T,self.p,self.rho,self.Fluid]:
            #Make the box not editable
            self.Fluid.SetEditable(False)
            #Bind events tp fire the chooser when text boxes are clicked on
            box.Bind(wx.EVT_LEFT_DOWN,self.UseChooser)
            box.SetToolTipString('Click on me to select the state')
        
        self.SetSizer(sizer)
        
        # create event class so that you can fire an event for when the state is changed
        # Other panels can await this event and do something when it happens
        import wx.lib.newevent as newevent
        self.StateUpdateEvent, self.EVT_STATE_UPDATE_EVENT = newevent.NewEvent()
        
    def GetState(self):
        """
        returns a :class:`CoolProp.State.State` instance from the given values
        in the panel
        """
        Fluid = str(self.Fluid.GetValue())
        T = float(self.T.GetValue())
        rho = float(self.rho.GetValue())
        return State(Fluid,dict(T = T, D = rho))
    
    def GetValue(self):
        return self.GetState()
    
    def SetValue(self, State_val):
        self.set_state(State_val.Fluid, dict(T = State_val.T, D = State_val.rho))
    
    def set_state(self, Fluid, **kwargs):
        """
        Fluid must not be unicode
        """
        if self._Fluid_fixed and not str(self.Fluid.GetValue()) == str(Fluid):
            import warnings
            warnings.warn('Could not set state since fluid is fixed')
            return

        #Create a state instance from the parameters passed in
        S  = State(str(Fluid), kwargs)

        #Load up the textboxes
        self.Fluid.SetValue(S.Fluid)
        self.T.SetValue(str(S.T))
        self.rho.SetValue(str(S.rho))
        self.p.SetValue(str(S.p))
    
    def UseChooser(self,event=None):
        """
        An event handler that runs the State Chooser dialog and sets the
        values back in the panel
        """
        #Values from the GUI
        Fluid = str(self.Fluid.GetValue())
        T = float(self.T.GetValue())
        rho = float(self.rho.GetValue())
        
        #Instantiate the chooser Dialog
        SCfrm=StateChooser(Fluid=Fluid,T=T,rho=rho,Fluid_fixed = self._Fluid_fixed)
        
        #If they clicked accept
        if wx.ID_OK == SCfrm.ShowModal():
            Fluid,T,p,rho=SCfrm.GetValues()
            #Set the GUI values
            self.Fluid.SetValue(str(Fluid))
            self.T.SetValue(str(T))
            self.p.SetValue(str(p))
            self.rho.SetValue(str(rho))
        SCfrm.Destroy()
        
        # post the update event, with arbitrary data attached
        wx.PostEvent(self, self.StateUpdateEvent())

class StateVariablesShimClass(object):
    """
    This class provides a shim that allows for state variables in the parametric
    table to set the value in the state panel
    
    It mirrors the functions in TextCtrl
    """
    
    def __init__(self, parent, val):
        self.val = val
        
    def SetValue(self, value):
        self.val = value
        
    def GetValue(self):
        return self.val
    
class StateInputsPanel(PDPanel):
    
    desc_map = dict(omega = ('Rotational speed [rad/s]','rad/s'),
                    inletState = ('The inlet state to the machine','-'),
                    discPratio = ('Pressure ratio (disc/suction)','-'),
                    discPressure = ('Discharge pressure [kPa]','kPa'),
                    discTsat = ('Discharge saturation temperature [K]','K'),
                    suctTsat = ('Suction saturation temperature [K]','K'),
                    suctDTsh = ('Suction superheat [K]','K'),
                    )
    
    def __init__(self, parent, config, **kwargs):
    
        PDPanel.__init__(self, parent, **kwargs)
        
        # The CoolProp State instance
        inletState = State(config['inletState']['Fluid'], dict(T = config['inletState']['T'], D = config['inletState']['rho']))
        
        # Create the sizers
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer_for_omega = wx.BoxSizer(wx.HORIZONTAL)
        sizer_for_inletState = wx.BoxSizer(wx.HORIZONTAL)
        sizer_for_discState = wx.FlexGridSizer(cols = 2)
        
        # The list for all the annotated objects
        self.annotated_values = []
        
        # Add the annotated omega to the list of objects
        self.annotated_values.append(AnnotatedValue('omega', config['omega'], *self.desc_map['omega'])) #*self.desc_map[key] unpacks into annotation, units
        
        # Construct annotated omega GUI entry
        AGO_omega = self.construct_items(self.annotated_values, sizer_for_omega)
        AGO_omega.GUI_location.SetToolTipString('If a motor curve is provided, this value will not be used')
        
        # Construct StatePanel
        self.SuctionStatePanel = StatePanel(self, CPState = inletState)
        
        self.Tsat = StateVariablesShimClass(self, 300)
        self.DTsh = StateVariablesShimClass(self, 11.1)
        
        # Construct the coupled annotated terms for the suction state
        CAGO_suctTsat = CoupledAnnotatedGUIObject(AnnotatedValue('suctTsat', self.Tsat, *self.desc_map['suctTsat']), self.Tsat, handler = self.parse_coupled_parametric_terms)
        CAGO_suctDTsh = CoupledAnnotatedGUIObject(AnnotatedValue('suctDTsh', self.DTsh, *self.desc_map['suctDTsh']), self.DTsh, handler = self.parse_coupled_parametric_terms)
        
        #Link required parameters
        CAGO_suctTsat.link_required_parameters([CAGO_suctDTsh])
        CAGO_suctDTsh.link_required_parameters([CAGO_suctTsat])
        
        #Construct the discharge state
        if 'pratio' in config['discharge']:
            pratio = config['discharge']['pratio']
            pressure = pratio * inletState.p
            Tsat = CP.Props('T','P',pressure,'Q',1,inletState.Fluid)
        elif 'pressure' in config['discharge']:
            pressure = config['discharge']['pressure']
            pratio = pressure / inletState.p
            Tsat = CP.Props('T','P',pressure,'Q',1,inletState.Fluid)
        else:
            raise ValueError('either pratio or pressure must be provided for discharge')
             
        disc_annotated_values = [
            AnnotatedValue('discPressure', pressure, *self.desc_map['discPressure']),
            AnnotatedValue('discPratio', pratio, *self.desc_map['discPratio']),
            AnnotatedValue('discTsat', Tsat, *self.desc_map['discTsat'])
            ]
        
        AGO_disc = self.construct_items(disc_annotated_values, sizer_for_discState)
        
        AGO_disc[0].GUI_location.Bind(wx.EVT_KILL_FOCUS,lambda event: self.OnChangeDischargeValue(event, 'pressure'))
        AGO_disc[1].GUI_location.Bind(wx.EVT_KILL_FOCUS,lambda event: self.OnChangeDischargeValue(event, 'pratio'))
        AGO_disc[2].GUI_location.Bind(wx.EVT_KILL_FOCUS,lambda event: self.OnChangeDischargeValue(event, 'Tsat'))
        
        self.main.register_GUI_objects([AGO_omega, CAGO_suctTsat, CAGO_suctDTsh] + AGO_disc)
        
        # Hack the discharge textctrl so that when they are updated using SetValue() they fire the EVT_KILL_FOCUS event
        def HackedSetValue(self, value):
            # Set the value without firing any events
            self.ChangeValue(value)
            # Create the event to be posted
            event = wx.PyCommandEvent(wx.EVT_KILL_FOCUS.typeId, self.GetId())
            # Post the event
            wx.PostEvent(self.GetEventHandler(), event)
            # And now process it
            self.ProcessPendingEvents()
            
        # Lambda to cut down code duplication
        get = lambda k: self.main.get_GUI_object(k).GUI_location
        
        for item in [get('discPressure'), get('discPratio'), get('discTsat')]:
            # Hook up the hacked bound SetValue method
            item.SetValue = types.MethodType( HackedSetValue, item )
        
        sizer.Add(HeaderStaticText(self,'Rotational Speed'), 0, wx.ALIGN_CENTER_HORIZONTAL)
        sizer.AddSpacer(5)
        sizer.Add(sizer_for_omega, 0, wx.ALIGN_CENTER_HORIZONTAL)
        sizer.AddSpacer(5)
        sizer.Add(HeaderStaticText(self,'Inlet State'), 0, wx.ALIGN_CENTER_HORIZONTAL)
        sizer.AddSpacer(5)
        sizer.Add(self.SuctionStatePanel, 0, wx.ALIGN_CENTER_HORIZONTAL)
        sizer.AddSpacer(5)
        sizer.Add(HeaderStaticText(self, 'Discharge State'), 0, wx.ALIGN_CENTER_HORIZONTAL)
        sizer.AddSpacer(5)
        sizer.Add(sizer_for_discState, 0, wx.ALIGN_CENTER_HORIZONTAL)
        
        self.SetSizer(sizer)
        sizer.Layout()
        
    def OnChangeDischargeValue(self, event = None, changed_parameter = ''):
        """ 
        Set the internal pressure variable when the value is changed in the TextCtrl
        """
        
        suction_state = self.SuctionStatePanel.GetState()
        psuction = suction_state.p
        Fluid = suction_state.Fluid
        
        if changed_parameter == 'pressure':
            pdisc = self.main.get_GUI_object_value('discPressure')
            pratio = pdisc / psuction
            Tsat = CP.Props('T', 'P', pdisc, 'Q', 1.0, Fluid)
            
        elif changed_parameter == 'pratio':
            pratio = self.main.get_GUI_object_value('discPratio')
            pdisc = psuction * pratio
            Tsat = CP.Props('T', 'P', pdisc, 'Q', 1.0, Fluid)
            
        elif changed_parameter == 'Tsat':
            Tsat = self.main.get_GUI_object_value('discTsat')
            pdisc = CP.Props('P', 'T', Tsat, 'Q', 1.0, Fluid)
            pratio = pdisc / psuction
            
        else:
            raise ValueError('Your parameter [{s:s}] is invalid for OnChangeDischargeValue.  Must be one of (pressure, pratio, Tsat)'.format(s = changed_parameter))
            
        # Set all the values again - we use ChangeValue manually to ensure that they 
        # don't emit change events which would result in infinite recursion
        self.main.get_GUI_object('discPressure').GUI_location.ChangeValue(str(pdisc))
        self.main.get_GUI_object('discPratio').GUI_location.ChangeValue(str(pratio))
        self.main.get_GUI_object('discTsat').GUI_location.ChangeValue(str(Tsat))
        
    def get_config_chunk(self):
        
        inletState = self.SuctionStatePanel.GetState()
        
        configdict = {}
        configdict['omega'] = self.main.get_GUI_object_value('omega')
        configdict['discharge'] = dict(pressure = self.main.get_GUI_object_value('discPressure'))
        configdict['inletState'] = dict(Fluid = inletState.Fluid,
                                        T = inletState.T,
                                        rho = inletState.rho)
        
        return configdict
        
    def get_script_chunks(self):
        """
        Get a string for the script file that will be run
        """
        inletState = self.SuctionStatePanel.GetState()
        discPressure = self.main.get_GUI_object_value('discPressure')
        omega = self.main.get_GUI_object_value('omega')
    
        return textwrap.dedent(
            """
            inletState = State.State("{Ref:s}", {{'T': {Ti:s}, 'P' : {pi:s} }})
    
            T2s = sim.guess_outlet_temp(inletState,{po:s})
            outletState = State.State("{Ref:s}", {{'T':T2s,'P':{po:s} }})
            
            # The rotational speed (over-written if motor map provided)
            sim.omega = {omega:s}
            """.format(Ref = inletState.Fluid,
                       Ti = str(inletState.T),
                       pi = str(inletState.p),
                       po = str(discPressure),
                       omega = str(omega)
                       )
            )
        
    def parse_coupled_parametric_terms(self, terms, keys, vals):
        
        # Map key to value from the parametric table
        val_map = { k:v for k,v in zip(keys,vals)}
        
        terms_keys = [t.key for t in terms]
        
        if 'suctDTsh' in terms_keys and 'suctTsat' in terms_keys:
            Tsat = val_map['suctTsat']
            DTsh = val_map['suctDTsh']
            
            Fluid = self.SuctionStatePanel.GetState().Fluid
            p = CP.Props('P','T',Tsat,'Q',1,Fluid)
            
            self.SuctionStatePanel.set_state(Fluid, **dict(T = Tsat+DTsh, P = p))
            
            # Update the discharge pressure ratio (Tsat disc and pdisc do not change)
            self.OnChangeDischargeValue(changed_parameter = 'pressure')
            
        else:
            raise NotImplementedException('This combination of coupled inlet state values is not supported')
    
class MotorCoeffsTable(wx.grid.Grid):
    
    def __init__(self, parent, values = None):
        """
        Parameters
        ----------
        parent : wx.window
            The parent of this checklist
        values : A 3-element list of lists for all the coeff (tau, eta, speed)
        """
        wx.grid.Grid.__init__(self, parent)
        
        # Make the grid the same shape as the data
        self.CreateGrid(20, 3) #Nrows, Ncolumns
        
        # Build the headers
        self.SetColLabelValue(0, 'Torque [N-m]')
        self.SetColLabelValue(1, 'Efficiency [-]')
        self.SetColLabelValue(2, 'Speed [rad/s]')
        
        # Set the entries in the grid
        self.update_from_configfile(values)
    
        # Bind the events
        self.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.OnCellRightClick)
        
    def OnCellRightClick(self, evt):
        
        # Make a menu
        menu = wx.Menu()
        
        #Build the entries
        menuitem1 = wx.MenuItem(menu, -1, 'Paste from clipboard (Excel format)')
        self.Bind(wx.EVT_MENU, self.OnPaste, menuitem1)
        menu.AppendItem(menuitem1)
        
        if menu.GetMenuItems():
            # Popup the menu.  If an item is selected then its handler
            # will be called before PopupMenu returns.
            self.PopupMenu(menu)
        
        menu.Destroy()
        
    def OnPaste(self, event):
        """
        Paste into the cells in the table
        """
        
        do = wx.TextDataObject()
        if wx.TheClipboard.Open():
            success = wx.TheClipboard.GetData(do)
            wx.TheClipboard.Close()

        data = do.GetText()
        rows = data.strip().replace('\r','').split('\n')
        rows = [row.split('\t') for row in rows]
        
        try:
            for row in rows:
                for el in row:
                    float(el)
            self.update_from_configfile(zip(*rows))
        except ValueError:
            dlg = wx.MessageDialog(None, "Unable to paste from clipboard - bad format")
            dlg.ShowModal()
            dlg.Close()
    
    def update_from_configfile(self, values):
        """
        
        Parameters
        ----------
        values : list of lists, with entries as floating point values
            The first entry is a list (or other iterable) of torque values
            
            The second entry is a list (or other iterable) of efficiency values
            
            The third entry is a list (or other iterable) of slip speed values
        """
        for i in range(self.GetNumberRows()):
            for j in range(self.GetNumberCols()):
                self.SetCellValue(i,0,'')
                
        for i,(tau,eff,speed) in enumerate(zip(*values)):
            if i == self.GetNumberRows():
                dlg  = wx.MessageDialog(None,"Sorry, too many values for the motor map table; truncating")
                dlg.ShowModal()
                dlg.Destroy()
                break
                
            # Values
            self.SetCellValue(i, 0, str(tau))
            self.SetCellValue(i, 1, str(eff))
            self.SetCellValue(i, 2, str(speed))
                
    def get_coeffs(self):
        """
        Get the list of lists of values that are used in the table
        """
        tau,eff,speed = [],[],[]
        for i in range(self.GetNumberRows()):
            _tau = self.GetCellValue(i, 0)
            _eff = self.GetCellValue(i, 1)
            _speed = self.GetCellValue(i, 2)
            
            #Check if row is empty
            if not _tau and not _eff and not _speed:
                continue
            
            #Convert to floating point if possible
            try:
                _tau = float(_tau)
                _eff = float(_eff)
                _speed = float(_speed)
            except ValueError:
                continue
            
            # Collect the values
            tau.append(_tau)
            eff.append(_eff)
            speed.append(_speed)
        return tau, eff, speed
        
class MotorChoices(wx.Choicebook):
    def __init__(self, parent):
        wx.Choicebook.__init__(self, parent, -1)
        
        self.pageconsteta=wx.Panel(self)
        self.AddPage(self.pageconsteta,'Constant efficiency')
        self.eta_motor_label, self.eta_motor = LabeledItem(self.pageconsteta, 
                                                           label="Motor Efficiency [-]",
                                                           value='0.9')
        sizer=wx.FlexGridSizer(cols = 2, hgap = 3, vgap = 3)
        sizer.AddMany([self.eta_motor_label, self.eta_motor])
        self.pageconsteta.SetSizer(sizer)
        
        self.pagemotormap=wx.Panel(self)
        self.AddPage(self.pagemotormap,'Motor map')
        self.MCT = MotorCoeffsTable(self.pagemotormap, values = [[1,2,3],[0.9,0.9,0.9],[307,307,307]])
        sizer = wx.FlexGridSizer(cols = 2, hgap = 3, vgap = 3)
        sizer.Add(self.MCT, 1, wx.EXPAND)
        self.pagemotormap.SetSizer(sizer)
        sizer.Layout()
        
if __name__=='__main__':    
    
    execfile('PDSimGUI.py')